#!/usr/bin/env python
# _*_ coding:utf-8 _*_
from openpyxl import load_workbook
from datetime import datetime
import time,requests
import smtplib
from email.mime.text import MIMEText
from email.header import Header
from email.mime.multipart import MIMEMultipart

#邮件的发送者和接受者
me = "309506489@qq.com"
# 多个收件人用list,单个收件人字符串
# accpter = ["1045033116@qq.com","309506489@qq.com","1031937206@qq.com"]
accpter = ['1045033116@qq.com','1031937206@qq.com','18612404428@163.com']
#每次循环等待间隔时间,默认60秒程序唤醒一次
waitTime = 60
#需要定时监测的时间点
timeList = ['10:30','11:30','13:45','14:30','14:45','14:55']
# timeList = ['20:07','20:09','20:11']
#基金代码
fundCode =['161725','000311','110022','161616','486001']

def runTaskRegularTime():
	while True:
		str_time_now = datetime.now().strftime('%H:%M')
		finish_time = datetime.now().strftime('%H')
		if int(finish_time)<15:#3点后结束
			if str(str_time_now) in timeList:#按照设定好的时间对比
			# if 1>0:#调试用
				matrix = doData()
				sendEmail(matrix)
				time.sleep(60)
			else:
				taskWait(str_time_now)
		else:
			print('闭市了\n')
			time.sleep(2)
			matrix = doData()
			sendEmail(matrix)
			write_excel(matrix)
			break

#把基金返回的数据存到二维数组里
def doData():
	matrix =[]
	for n in range(len(fundCode)):
		list = inquiryRate(fundCode[n])
		matrix.append(list)
	# print(matrix[0])
	return matrix

def taskWait(currentTime):
	print('current time：' + str(currentTime))
	print('process will wakes up every %s seconds because it does not arrive at the specified time' % (waitTime))
	print ('')
	print ('')
	time.sleep(waitTime)

#调接口查询
def inquiryRate(shourNumber):
	try:
		url = (" http://fundmobapi.eastmoney.com/FundMApi/FundVarietieValuationDetail.ashx?version=5.3.0&plat=Android&appType=ttjj&FCODE=%s&deviceid=e19655bcbef4c4b362d908f8bcbdd67f||946596830144568&product=EFund&MobileKey=e19655bcbef4c4b362d908f8bcbdd67f||946596830144568"%shourNumber)
		r = requests.get(url, params=None)
		# print(r.json()["Expansion"])
		today = r.json()["Expansion"]["JZRQ"]
		fcode = r.json()["Expansion"]["FCODE"]
		shortname = r.json()["Expansion"]["SHORTNAME"]
		yn = r.json()["Expansion"]["DWJZ"]
		tn = r.json()["Expansion"]["GZ"]
		upordown = r.json()["Expansion"]["GSZZL"]
		# print("_____________________")
		ex_content = [today, fcode, shortname, yn, tn, upordown]

		return ex_content
	except Exception as e:
		print('接口有问题'+str(e))
	return


def sendEmail(list):
	msg = MIMEMultipart()
	msg['from'] = me
	msg['to'] = ','.join(accpter)
	msg['subject'] = Header("%s基金数据"%datetime.now().strftime('%H:%M'), 'utf-8')
	list1= []
	for n in range(len(list)):
		lista = list[n]
		mailBody = '    '.join(lista)
		list1.append(mailBody)
	mailBody = '\n'.join(list1)
	print(mailBody)
	puretext = MIMEText(mailBody, 'plain', 'utf-8')
	msg.attach(puretext)
	try:
		server = smtplib.SMTP_SSL('smtp.qq.com', '465')
		server.login(me, "lhiawwpunkcrbjeg")
		server.sendmail(me, accpter, msg.as_string())
		server.quit()

	except Exception as e:
		print(str(e))
	return

#读取项目的excel，修改添加数据后保存
def write_excel(list):
	wb = load_workbook('myfundStatistics.xlsx')
	sheet = wb['Sheet1']
	n_coordinate = 0
	for row in sheet.iter_rows():#找到空白行号
		for cell in row:
			# print(cell.coordinate, cell.value)
			if cell.value is not None:
				n_coordinate = cell.row+1
	# print(n_coordinate)
	for n in range(len(list)):#通过遍历给cell赋值
		for m in range(len(list[0])):
			sheet.cell(n_coordinate+n,m+1).value=list[n][m]
	wb.save('myfundStatistics.xlsx')

if __name__ == '__main__':
	runTaskRegularTime()